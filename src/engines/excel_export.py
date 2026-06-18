import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import logging

from src.database.connection import get_connection
from src.database.queries import get_dashboard_stats

# Define Corporate Color Palette (Navy / Charcoal / Muted Gray)
NAVY_HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
GRAY_ACCENT_FILL = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
WHITE_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
REGULAR_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F497D")

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

def sanitize_formula_injection(val):
    """Prepend a single quote if string starts with formula indicators to block CSV Injection."""
    if isinstance(val, str) and val.startswith(('=', '+', '-', '@')):
        return f"'{val}"
    return val

def export_all_data_to_excel(output_path: str) -> tuple:
    """
    Compile database records and export them to a professional styled Excel workbook.
    Returns (success: bool, error_message: str or None)
    """
    try:
        # Create output directory
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Initialize workbook
        wb = openpyxl.Workbook()
        
        # 1. Sheet: Dashboard Summary
        ws_dash = wb.active
        ws_dash.title = "Dashboard Summary"
        build_dashboard_sheet(ws_dash)
        
        # 2. Sheet: Student Directory
        ws_stud = wb.create_sheet(title="Student Directory")
        build_student_sheet(ws_stud)
        
        # 3. Sheet: Lead Pipeline
        ws_lead = wb.create_sheet(title="Lead Pipeline")
        build_lead_sheet(ws_lead)
        
        # 4. Sheet: Fee Collection & Dues
        ws_fee = wb.create_sheet(title="Fee Ledger")
        build_fee_sheet(ws_fee)
        
        # Save workbook
        wb.save(output_path)
        logging.info(f"Excel summary successfully exported to: {output_path}")
        return True, None
    except Exception as e:
        logging.error(f"Failed to export Excel report: {e}")
        return False, str(e)

def autofit_columns_and_freeze(ws, freeze_cell="A5"):
    """Format gridlines, autofit column widths based on content, and freeze panes."""
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = freeze_cell
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        # Set buffer width
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

def build_dashboard_sheet(ws):
    """Compile high-level stats and insert visual summary elements."""
    stats = get_dashboard_stats()
    
    # Gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws.merge_cells("A1:D1")
    ws["A1"] = "AcademyOS - Executive Summary"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    
    ws["A2"] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10)
    
    # Header styling
    ws.row_dimensions[4].height = 25
    headers = ["Metric Description", "Value", "", "Target Allocation Chart"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        if h:
            cell.fill = NAVY_HEADER_FILL
            cell.font = WHITE_FONT
            cell.alignment = Alignment(horizontal="center" if i > 1 else "left", vertical="center")
            
    # Metrics
    metrics = [
        ("Total Active Students Enrolled", stats["active_students"]),
        ("Total Leads in Sales Pipeline", stats["active_leads"]),
        ("Total Offered Courses", stats["total_courses"]),
        ("Total Net Receivable Fees ($)", stats["total_receivable"]),
        ("Total Collected Fees ($)", stats["total_collected"]),
        ("Total Outstanding Dues ($)", stats["pending_receivable"])
    ]
    
    for row_idx, (desc, val) in enumerate(metrics, start=5):
        c1 = ws.cell(row=row_idx, column=1, value=desc)
        c2 = ws.cell(row=row_idx, column=2, value=val)
        
        # Border & fonts
        c1.font = REGULAR_FONT
        c1.border = THIN_BORDER
        c2.font = BOLD_FONT
        c2.border = THIN_BORDER
        
        if "$" in desc or "Receivable" in desc or "Collected" in desc or "Outstanding" in desc:
            c2.number_format = '$#,##0.00'
            c2.alignment = Alignment(horizontal="right")
        else:
            c2.alignment = Alignment(horizontal="center")
            
    # Draw simple charts (Dues vs Collection)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Financial Summary (Collection vs Outstanding)"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Category"
    
    # Reference metrics rows: rows 8 (receivable), 9 (collected), 10 (dues)
    data = Reference(ws, min_col=2, min_row=7, max_row=10)
    cats = Reference(ws, min_col=1, min_row=7, max_row=10)
    
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.legend = None  # No legend needed for single category chart
    
    ws.add_chart(chart, "D5")
    
    # Adjust widths manually for dashboard
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

def build_student_sheet(ws):
    """Fetch students directory, construct grid view, and apply conditional colors."""
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Student Enrollment Directory"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    
    # Setup headers
    headers = ["ID", "First Name", "Last Name", "Email", "Phone", "DOB", "Admission Date", "Status"]
    ws.row_dimensions[4].height = 24
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=h)
        cell.fill = NAVY_HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center" if idx in (1, 6, 7, 8) else "left", vertical="center")
        
    # Query details
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, first_name, last_name, email, phone, dob, admission_date, status FROM students ORDER BY id ASC;")
    rows = cursor.fetchall()
    
    for row_idx, row in enumerate(rows, start=5):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=sanitize_formula_injection(val))
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            if col_idx in (1, 6, 7, 8):
                cell.alignment = Alignment(horizontal="center")
                
    cursor.close()
    
    # Apply Auto filters
    max_row = max(4, len(rows) + 4)
    ws.auto_filter.ref = f"A4:H{max_row}"
    
    # Freeze panes below headers
    autofit_columns_and_freeze(ws, "A5")

def build_lead_sheet(ws):
    """Fetch leads information and apply layout formatting."""
    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "Enquiry and Lead Tracking Pipeline"
    ws["A1"].font = TITLE_FONT
    
    headers = ["ID", "First Name", "Last Name", "Phone", "Email", "Source", "Status", "Follow-up Date", "Remarks"]
    ws.row_dimensions[4].height = 24
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=h)
        cell.fill = NAVY_HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center" if idx in (1, 7, 8) else "left", vertical="center")
        
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, first_name, last_name, phone, email, source, status, follow_up_date, remarks FROM leads ORDER BY created_at DESC;")
    rows = cursor.fetchall()
    
    for row_idx, row in enumerate(rows, start=5):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=sanitize_formula_injection(val))
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            if col_idx in (1, 7, 8):
                cell.alignment = Alignment(horizontal="center")
                
    cursor.close()
    
    max_row = max(4, len(rows) + 4)
    ws.auto_filter.ref = f"A4:I{max_row}"
    autofit_columns_and_freeze(ws, "A5")

def build_fee_sheet(ws):
    """Export the list of students with active course enrollments, payments, and dues."""
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Fee Ledger & Outstanding Balances"
    ws["A1"].font = TITLE_FONT
    
    headers = ["Student ID", "Student Name", "Course Enrolled", "Net Receivable Fee", "Amount Collected", "Pending Balance", "Enrollment Status", "Last Payment Date"]
    ws.row_dimensions[4].height = 24
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=h)
        cell.fill = NAVY_HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center" if idx in (1, 7, 8) else "left", vertical="center")
        
    conn = get_connection()
    cursor = conn.cursor()
    
    # Query fee collection aggregates per enrollment
    cursor.execute(
        """
        SELECT 
            s.id as student_id,
            s.first_name || ' ' || s.last_name as student_name,
            c.name as course_name,
            e.id as enrollment_id,
            e.net_fee,
            COALESCE((SELECT SUM(amount_paid) FROM fee_payments WHERE enrollment_id = e.id), 0.0) as total_paid,
            e.status as enrollment_status,
            (SELECT MAX(payment_date) FROM fee_payments WHERE enrollment_id = e.id) as last_payment_date
        FROM enrollments e
        JOIN students s ON e.student_id = s.id
        JOIN courses c ON e.course_id = c.id
        ORDER BY student_name ASC;
        """
    )
    rows = cursor.fetchall()
    
    for row_idx, row in enumerate(rows, start=5):
        ws.row_dimensions[row_idx].height = 20
        
        student_id = row["student_id"]
        student_name = row["student_name"]
        course_name = row["course_name"]
        net_fee = row["net_fee"]
        total_paid = row["total_paid"]
        due_amount = max(0.0, net_fee - total_paid)
        status = row["enrollment_status"]
        last_pay = row["last_payment_date"] or "N/A"
        
        # Write cells
        c_id = ws.cell(row=row_idx, column=1, value=student_id)
        c_name = ws.cell(row=row_idx, column=2, value=sanitize_formula_injection(student_name))
        c_course = ws.cell(row=row_idx, column=3, value=sanitize_formula_injection(course_name))
        c_net = ws.cell(row=row_idx, column=4, value=net_fee)
        c_paid = ws.cell(row=row_idx, column=5, value=total_paid)
        c_due = ws.cell(row=row_idx, column=6, value=due_amount)
        c_status = ws.cell(row=row_idx, column=7, value=status)
        c_date = ws.cell(row=row_idx, column=8, value=last_pay)
        
        for cell in (c_id, c_name, c_course, c_net, c_paid, c_due, c_status, c_date):
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            
        c_id.alignment = Alignment(horizontal="center")
        c_status.alignment = Alignment(horizontal="center")
        c_date.alignment = Alignment(horizontal="center")
        
        c_net.number_format = '$#,##0.00'
        c_paid.number_format = '$#,##0.00'
        c_due.number_format = '$#,##0.00'
        
        c_net.alignment = Alignment(horizontal="right")
        c_paid.alignment = Alignment(horizontal="right")
        c_due.alignment = Alignment(horizontal="right")
        
    cursor.close()
    
    # Conditional formatting: highlight pending balances > 0 with light red fill
    max_row = max(4, len(rows) + 4)
    ws.auto_filter.ref = f"A4:H{max_row}"
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(name="Calibri", size=11, color="9C0006")
    
    ws.conditional_formatting.add(
        f"F5:F{max_row}",
        CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font)
    )
    
    autofit_columns_and_freeze(ws, "A5")
